import tkinter as tk
from tkinter import ttk, messagebox
from tkinter.filedialog import asksaveasfilename
import threading
from openpyxl import Workbook
from playwright.sync_api import sync_playwright
import random
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class GlassdoorScraperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Glassdoor Job Scraper")
        self.root.geometry("520x320")  
        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.grid(row=0, column=0, sticky="nsew")
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        # Job Title
        ttk.Label(main_frame, text="Job Title:", font=("Segoe UI", 12)).grid(row=0, column=0, sticky=tk.W, padx=5, pady=10)
        self.job_entry = ttk.Entry(main_frame, width=32, font=("Segoe UI", 12))
        self.job_entry.grid(row=0, column=1, padx=5, pady=10)
        # Location
        ttk.Label(main_frame, text="Location:", font=("Segoe UI", 12)).grid(row=1, column=0, sticky=tk.W, padx=5, pady=10)
        self.location_entry = ttk.Entry(main_frame, width=32, font=("Segoe UI", 12))
        self.location_entry.grid(row=1, column=1, padx=5, pady=10)
        # Number of jobs
        ttk.Label(main_frame, text="Number of Jobs:", font=("Segoe UI", 12)).grid(row=2, column=0, sticky=tk.W, padx=5, pady=10)
        self.num_entry = ttk.Entry(main_frame, width=10, font=("Segoe UI", 12))
        self.num_entry.grid(row=2, column=1, padx=5, pady=10, sticky=tk.W)
        # Add helpful instructions
        ttk.Label(main_frame, text="1. Enter your search details.\n2. Click Start Scraping.\n3. Log in to Glassdoor in the browser window.\n4. Data will be saved to Excel.", font=("Segoe UI", 10), foreground="#555").grid(row=3, column=0, columnspan=2, pady=5)
        # Start button (move to bottom, full width)
        self.start_btn = ttk.Button(main_frame, text="Start Scraping", command=self.start_scraping_thread)
        self.start_btn.grid(row=4, column=0, columnspan=2, pady=20, sticky="ew")

    def start_scraping_thread(self):
        threading.Thread(target=self.scrape_jobs, daemon=True).start()

    def scrape_jobs(self):
        job = self.job_entry.get().strip()
        location = self.location_entry.get().strip()
        try:
            num_jobs = int(self.num_entry.get().strip())
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number of jobs.")
            return
        self.start_btn.config(state=tk.DISABLED)
        try:
            data = self.run_playwright(job, location, num_jobs)
            if not data:
                messagebox.showinfo("No Data", "No jobs were scraped.")
                return
            save_path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if not save_path:
                return
            self.save_to_excel(data, save_path)
            messagebox.showinfo("Success", f"Saved {len(data)} jobs to {save_path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            self.start_btn.config(state=tk.NORMAL)

    def wait_like_human(self, min_sec=1, max_sec=2.5):
        time.sleep(random.uniform(min_sec, max_sec))

    def run_playwright(self, job, location, num_jobs):
        data = []
        import os
        user_data_dir = os.path.join(os.path.dirname(__file__), "glassdoor_profile")
        chrome_user_agent = (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/123.0.0.0 Safari/537.36"
        )
        with sync_playwright() as p:
            browser = p.chromium.launch_persistent_context(
                user_data_dir,
                headless=False,
                args=["--start-maximized"],
            )
            context = browser.new_page()
            context.set_extra_http_headers({"User-Agent": chrome_user_agent})
            page = context
            page.goto("https://www.glassdoor.com/")
            self.wait_like_human(3.5, 4.5)  # Wait for Cloudflare check
            login_needed = False
            if page.url.startswith("https://www.glassdoor.com/profile/login_input.htm") or "login" in page.url:
                login_needed = True
            else:
                try:
                    if page.query_selector('a[data-test="sign-in-link"]'):
                        login_needed = True
                except Exception:
                    pass
            if login_needed:
                messagebox.showinfo("Manual Login Required", "Please log in to Glassdoor and solve any captcha in the opened browser window.\n\nAfter you see your profile or the homepage, close the browser window to continue.")
                while True:
                    if len(browser.pages) == 0:
                        break
                    self.wait_like_human(0.5, 1.2)
                browser = p.chromium.launch_persistent_context(
                    user_data_dir,
                    headless=False,
                    args=["--start-maximized"],
                )
                context = browser.new_page()
                context.set_extra_http_headers({"User-Agent": chrome_user_agent})
                page = context
            # Now proceed to search
            page.goto("https://www.glassdoor.com/Job/jobs.htm")
            self.wait_like_human(2, 3)
            # Try multiple selectors for robustness
            job_selectors = [
                'input[placeholder="Find your perfect job"]',
                'input[data-test="search-job-keyword-input-internal"]',
                'input#searchJobKeyword',
                'input[placeholder*="Job title"]',
                'input[aria-label*="Search Keyword"]',
                'input[name="keyword"]',
            ]
            location_selectors = [
                'input[placeholder*="City, state, zipcode"]',
                'input[id^="searchBar-location"]',
                'input[data-test="search-job-location-input-internal"]',
                'input#searchJobLocation',
                'input[placeholder*="Location"]',
                'input[aria-label*="Search Location"]',
                'input[name="location"]',
            ]
            job_input = None
            location_input = None
            # Fill job input
            for sel in job_selectors:
                try:
                    page.wait_for_selector(sel, timeout=3000)
                    job_input = page.query_selector(sel)
                    if job_input:
                        job_input.click()
                        job_input.fill("")
                        self.wait_like_human(0.3, 0.7)
                        job_input.type(job)
                        self.wait_like_human(0.3, 0.7)
                        job_input.press("Enter")
                        self.wait_like_human(0.5, 1)
                        break
                except Exception:
                    continue
            # Fill location input
            for sel in location_selectors:
                try:
                    page.wait_for_selector(sel, timeout=3000)
                    location_input = page.query_selector(sel)
                    if location_input:
                        location_input.click()
                        location_input.fill("")
                        self.wait_like_human(0.3, 0.7)
                        location_input.type(location)
                        self.wait_like_human(0.3, 0.7)
                        location_input.press("Enter")
                        page.wait_for_load_state('networkidle', timeout=20000)
                        self.wait_like_human(2, 3)
                        break
                except Exception as e:
                    print(f"[DEBUG] Could not fill job/location fields: {e}")
            # After job, click location to focus it (use robust selector) and select correct dropdown option
            try:
                page.wait_for_selector('input[id^="searchBar-location"]', timeout=3000)
                location_input = page.query_selector('input[id^="searchBar-location"]')
                if location_input:
                    location_input.click()
                    location_input.fill("")
                    self.wait_like_human(0.3, 0.7)
                    # Type location char by char
                    for char in location:
                        location_input.type(char)
                        self.wait_like_human(0.05, 0.15)
                    self.wait_like_human(1, 1.5)  # Wait for dropdown
                    # Find dropdown options and select the correct one
                    dropdown_options = page.query_selector_all('li[role="option"], div[role="option"]')
                    found = False
                    for option in dropdown_options:
                        try:
                            option_text = option.inner_text().strip().lower()
                            if location.lower() in option_text:
                                option.click()
                                found = True
                                self.wait_like_human(0.3, 0.7)
                                break
                        except Exception:
                            continue
                    if not found:
                        # fallback: ArrowDown+Enter
                        location_input.press("ArrowDown")
                        self.wait_like_human(0.2, 0.4)
                        location_input.press("Enter")
                        self.wait_like_human(0.3, 0.7)
                    # After selecting location, press Enter to trigger reload
                    location_input.press("Enter")
                    self.wait_like_human(2, 3)  # Wait for page reload
                    page.wait_for_load_state('networkidle', timeout=20000)
                    self.wait_like_human(1, 2)
            except Exception:
                print(f"[DEBUG] Could not fill location field with robust selector")
            count = 0
            while count < num_jobs:
                # Scroll slowly through the page like a human
                for y in range(0, 2000, 200):
                    page.evaluate(f"window.scrollBy({{ top: 200, behavior: 'smooth' }})")
                    self.wait_like_human(0.5, 1.2)
                job_cards = page.query_selector_all('li[data-test="jobListing"]')
                for card in job_cards:
                    if count >= num_jobs:
                        break
                    try:
                        self.wait_like_human(1, 2.5)
                        # Click the job card to open the detailed view
                        card.click()
                        self.wait_like_human(1.5, 2.5)
                        # Wait for job title in detail view
                        try:
                            page.wait_for_selector('h1[id^="jd-job-title-"]', timeout=5000)
                        except Exception:
                            pass
                        # Extract job title and company name from detail view
                        title = ""
                        company = ""
                        try:
                            title_el = page.query_selector('h1[id^="jd-job-title-"]')
                            if title_el:
                                title = title_el.inner_text().strip()
                        except Exception:
                            pass
                        # Try multiple selectors for company name
                        try:
                            company_el = page.query_selector('div[data-test="employerName"]')
                            if not company_el:
                                company_el = page.query_selector('span[data-test="employerName"]')
                            if not company_el:
                                company_el = page.query_selector('a[data-test="employerName"]')
                            if not company_el:
                                # Try the profile container link with id pattern
                                company_el = page.query_selector('a[id^="3"][class*="EmployerProfile_profileContainer"]')
                            if not company_el:
                                # Fallback to any profile container
                                company_el = page.query_selector('a[class*="EmployerProfile_profileContainer"]')
                            if company_el:
                                company_text = company_el.inner_text().strip()
                                # Clean up the company name
                                import re
                                # First remove rating if present (e.g., "4.5★")
                                company_text = re.sub(r'\s*\d+(\.\d+)?★.*$', '', company_text)
                                # Then remove "Logo" and anything before it
                                if 'Logo' in company_text:
                                    parts = company_text.split('Logo')
                                    # Take the part after "Logo"
                                    company_text = parts[-1].strip()
                                # Remove any duplicate spaces and clean up
                                company = re.sub(r'\s+', ' ', company_text).strip()
                        except Exception as e:
                            print(f"[DEBUG] Error extracting company name: {e}")
                            pass
                        # Fallback to card if detail view fails
                        if not title:
                            try:
                                title_el = card.query_selector('a[data-test="job-link"]')
                                title = title_el.inner_text().strip() if title_el else ""
                            except Exception:
                                title = ""
                        if not company:
                            try:
                                company_el = card.query_selector('div[data-test="jobListing-company-name"]')
                                if not company_el:
                                    company_el = card.query_selector('span[data-test="jobListing-company-name"]')
                                if company_el:
                                    company = company_el.inner_text().strip()
                            except Exception:
                                company = ""
                        # Get direct link
                        link = ""
                        try:
                            link_el = card.query_selector('a[data-test="job-link"]')
                            if link_el:
                                link = link_el.get_attribute('href')
                                if link and not link.startswith('http'):
                                    link = "https://www.glassdoor.com" + link
                        except Exception:
                            link = ""
                        data.append({
                            'Company Name': company,
                            'Job Title': title,
                            'Direct Link': link
                        })
                        count += 1
                    except Exception:
                        continue
                if count < num_jobs:
                    self.wait_like_human(1.2, 3)
                    next_btn = page.query_selector('button[data-test="pagination-next"]')
                    if next_btn and not next_btn.is_disabled():
                        next_btn.click()
                        self.wait_like_human(1.5, 3)
                    else:
                        break
            browser.close()
        return data

    def save_to_excel(self, data, file_path):
        if not data:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Listings"
        # Write header
        headers = list(data[0].keys())
        ws.append(headers)
        for header in headers:
            ws.column_dimensions[get_column_letter(headers.index(header) + 1)].width = 25  # Set column width
            
        # Get the column index for Direct Link
        link_col_idx = headers.index('Direct Link') + 1
        
        # Write data and create hyperlinks
        for row_idx, row_data in enumerate(data, start=2):  # start=2 because header is row 1
            for col_idx, value in enumerate(row_data.values(), start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx == link_col_idx and value:  # If this is the Direct Link column
                    cell.hyperlink = value  # Set the hyperlink
                    cell.value = "Click here"  # Change display text
                    cell.font = Font(color="0000FF", underline="single")  # Make it look like a link
                else:
                    cell.value = value
                    
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
        # Save workbook
        wb.save(file_path)
        wb.close()

def on_closing():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        root.destroy()

root = tk.Tk()
app = GlassdoorScraperApp(root)
root.protocol("WM_DELETE_WINDOW", on_closing)
root.mainloop()
